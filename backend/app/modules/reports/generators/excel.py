"""Base Excel generation utilities using openpyxl."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Reusable styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HIGHLIGHT_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")


def create_workbook() -> Workbook:
    """Create a new workbook with default settings."""
    wb = Workbook()
    return wb


def add_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: Sequence[Sequence[Any]],
    *,
    highlight_rows: set[int] | None = None,
) -> None:
    """Add a formatted sheet to the workbook.

    Args:
        wb: Target workbook.
        title: Sheet title.
        headers: Column header names.
        rows: Data rows (list of sequences).
        highlight_rows: Row indices (0-based) to highlight in orange.
    """
    if wb.active and wb.active.title == "Sheet":
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    # Write data
    for row_idx, row_data in enumerate(rows):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx + 2, column=col_idx)
            if isinstance(value, Decimal):
                cell.value = float(value)
                cell.number_format = "#,##0.00"
            elif isinstance(value, datetime):
                cell.value = value
                cell.number_format = "YYYY-MM-DD HH:MM"
            elif isinstance(value, date):
                cell.value = value
                cell.number_format = "YYYY-MM-DD"
            else:
                cell.value = value

            if highlight_rows and row_idx in highlight_rows:
                cell.fill = HIGHLIGHT_FILL

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                val_len = len(str(row_data[col_idx - 1] or ""))
                if val_len > max_len:
                    max_len = val_len
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize a workbook to bytes for HTTP response."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
